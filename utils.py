import pickle
from openpyxl import Workbook
from openpyxl import load_workbook
import discretization as dis
import math

def calcola_distanza(differenze):
    # differenze[0] contiene la differenza delle coordinate x
    # differenze[1] contiene la differenza delle coordinate y
    distanza = math.sqrt(differenze[0]**2 + differenze[1]**2)
    return distanza

def writeFinalResultsXls(fileName, solutions, gray_codes, fitnesses, evaluations, distances):
    """Function that writes the final results on an excel file"""
    wb = Workbook()
    ws1 = wb.active

    ws1.cell(column=1, row=1, value="Chromosome")
    ws1.cell(column=2, row=1, value="Gray code")
    ws1.cell(column=3, row=1, value="Fitness Value")
    ws1.cell(column=4, row=1, value="Evaluations to obtain best")
    ws1.cell(column=5, row=1, value="Hamming distance")

    row = 2
    for g in solutions:
        ws1.cell(column=1, row=row, value=g)
        row += 1

    row = 2
    for g in gray_codes:
        ws1.cell(column=2, row=row, value=g)
        row += 1

    row = 2
    for g in fitnesses:
        ws1.cell(column=3, row=row, value=g)
        row += 1

    row = 2
    for g in evaluations:
        ws1.cell(column=4, row=row, value=g)
        row += 1

    row = 2
    for g in distances:
        ws1.cell(column=5, row=row, value=str(g))
        row += 1

    wb.save(filename = fileName)

def writeChromosomeEvolutionXls(fileName, chromosome_evolution):
    """Function that writes the chromosome evolution object on an excel file"""
    wb = Workbook()
    ws1 = wb.active

    for i in range(1, len(chromosome_evolution[0])+1):
        ws1.cell(column=i, row=1, value="Chromosome "+str(i))


    row = 2
    for l_sup in chromosome_evolution:
        col=1
        for c in l_sup:
            ws1.cell(column=col, row=row, value=str(c))
            col+=1
        row += 1

    wb.save(filename = fileName)

def writeBestsXls(fileName, bests):
    """Function that writes the bests object on an excel file"""
    wb = Workbook()
    ws1 = wb.active

    ws1.cell(column=1, row=1, value="Best Chromosome")
    ws1.cell(column=2, row=1, value="Fitness")
    ws1.cell(column=3, row=1, value="Gray code")

    row = 2
    for l_sup in bests:
        col=1
        for c in l_sup:
            ws1.cell(column=col, row=row, value=str(c))
            col+=1
        row += 1

    wb.save(filename = fileName)


def readXls(filename, col_name, flagFloat=True, ind_min=None, ind_max=None):
    """Function that reads a value from an excel file"""
    wb = load_workbook(filename=filename)
    sheet = wb['Sheet']
    if ind_min==None and ind_max==None:
        fitness_col = sheet[col_name][1:]
    elif ind_min==None:
        fitness_col = sheet[col_name][1:ind_max+1]
    elif ind_max==None:
        fitness_col = sheet[col_name][ind_min:]
    else:
        fitness_col = sheet[col_name][ind_min:ind_max+1]

    l=[]
    for cell in fitness_col:
        v=cell.value
        if flagFloat==True:
            v=float(v)
        l.append(v)
    return l


def hamming_distance(chr1, chr2, lower_bounds, upper_bounds, num_bit_code, dim, base):
    """Function that computes the hamming distance between two individuals"""
    fen1=dis.convertFromBinToFloat(chr1, lower_bounds, upper_bounds, num_bit_code, dim, base)
    fen2=dis.convertFromBinToFloat(chr2, lower_bounds, upper_bounds, num_bit_code, dim, base)
    step = [(x - y) / (base ** num_bit_code - 1) for x,y in zip(upper_bounds, lower_bounds)]
    dist=[x-y for x, y in zip(fen1,fen2)]
    dist=[x/y for x, y in zip(dist,step)]
    distanza=calcola_distanza(dist)
    return distanza


def computeHammingDistance(colgray, problem):
    """Function that computes the Hamming distance between an individual and the optimal individual"""
    best_sols=problem.getOptimum()[1]
    distances=[]
    for sol in best_sols:
            best_opt=''
            for s in sol:
                best_opt+=s
            distances.append(hamming_distance(colgray, best_opt))
    return min(distances)

